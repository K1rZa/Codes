from flask import Flask, request, render_template, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'

def clear_upload_folder():
    """Удаляет все файлы в папке uploads."""
    for filename in os.listdir(app.config['UPLOAD_FOLDER']):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f"Ошибка при удалении файла {file_path}: {e}")

def compare_excel_files(user_file_path, reference_file_path, task_number, variant_number):
    df_user = pd.read_excel(user_file_path)
    df_reference = pd.read_excel(reference_file_path)

    wb_user = load_workbook(user_file_path)
    ws_user = wb_user.active

    user_name = ws_user['A1'].value
    group_number = ws_user['B1'].value

    headers = df_reference.columns.tolist()
    correct_count = 0  
    total_count = 0   

    for index, row_user in df_user.iterrows():
        if index >= len(df_reference):
            break  # 

        row_reference = df_reference.iloc[index]

        for col_index, header in enumerate(headers):
            cell = ws_user.cell(row=index + 2, column=col_index + 1)  

            cell_value_user = row_user.get(header)  
            cell_value_reference = row_reference.get(header)

            if cell_value_reference is not None:
                total_count += 1

            if cell_value_user is not None and cell_value_reference is not None:
                if cell_value_user != cell_value_reference:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.value = cell_value_user
                else:
                    cell.value = cell_value_user
                    correct_count += 1
            else:
                cell.value = cell_value_user if cell_value_user is not None else cell_value_reference

    percentage_correct = ((correct_count + 1) / (total_count - 1) * 100) if total_count > 0 else 0

    ws_user['A1'].value = f"{percentage_correct:.2f}%"
    ws_user['B1'].value = None

    wb_user.save(user_file_path)

    new_filename = f"{user_name.replace(' ', '_')}_{group_number}_Вариант_{variant_number}_Задание_{task_number}.xlsx"
    new_file_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)

    os.rename(user_file_path, new_file_path)

    return new_file_path

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        clear_upload_folder()

        task_number = request.form['task_number']
        variant_number = request.form['variant_number']
        file1 = request.files['file1']
        
        if file1:
            file1_path = os.path.join(app.config['UPLOAD_FOLDER'], file1.filename)
            file1.save(file1_path)

            file2_path = os.path.join('correct', str(variant_number), f'task{task_number}_variant{variant_number}.xlsx')

            new_file_path = compare_excel_files(file1_path, file2_path, task_number, variant_number)

            return send_file(new_file_path, as_attachment=True)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)

